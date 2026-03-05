"""
Helper script to create a sample Excel input file for the Student Grade Calculator.
This script creates students_input.xlsx with sample student data.

Requirements: pip install openpyxl
"""

from openpyxl import Workbook

def create_sample_input_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Headers
    ws['A1'] = "Student Name"
    ws['B1'] = "Score"
    
    # Sample data
    students = [
        ("Alice Johnson", 95),
        ("Bob Smith", 82),
        ("Charlie Brown", 67),
        ("Diana Prince", 45),
        ("Eve Williams", 88),
        ("Frank Miller", 73),
        ("Grace Lee", 91),
        ("Henry Davis", 58),
        ("Iris Chen", 76),
        ("Jack Wilson", 84),
    ]
    
    # Add student data
    for idx, (name, score) in enumerate(students, start=2):
        ws[f'A{idx}'] = name
        ws[f'B{idx}'] = score
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    
    # Save file
    filename = "students_input.xlsx"
    wb.save(filename)
    print(f"✓ Created {filename} with {len(students)} students")
    print("\nSample data:")
    for name, score in students[:5]:
        print(f"  - {name}: {score}")
    print(f"  ... and {len(students) - 5} more")

if __name__ == "__main__":
    create_sample_input_file()
